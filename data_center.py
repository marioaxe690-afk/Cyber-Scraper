import requests
from bs4 import BeautifulSoup
import pandas as pd
import time
import urllib3
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

# 禁用 SSL 警告
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# 战术伪装：User-Agent 字典
HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
}


def beautify_excel(file_path, column_widths):
    """
    通用 Excel 美化函数
    :param file_path: Excel 文件路径
    :param column_widths: 字典，键为列名，值为列宽
    """
    wb = load_workbook(file_path)
    ws = wb.active

    # 冻结首行
    ws.freeze_panes = 'A2'

    # 设置表头样式（加粗）
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # 设置列宽和自动换行
    for col_idx, cell in enumerate(ws[1], start=1):
        col_letter = cell.column_letter
        col_name = cell.value

        # 设置列宽
        if col_name in column_widths:
            ws.column_dimensions[col_letter].width = column_widths[col_name]
        else:
            ws.column_dimensions[col_letter].width = 15  # 默认宽度

    # 设置所有数据单元格自动换行
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    wb.save(file_path)
    print(f"[✓] Excel 美化完成: {file_path}")



def academic_radar():
    """[1] 学术前沿雷达 - 抓取 arXiv 最新 CS 论文"""
    print("\n[*] 启动学术前沿雷达...")

    url = "https://arxiv.org/list/cs.AI/recent"

    try:
        response = requests.get(url, headers=HEADERS, timeout=10, verify=False, proxies={"http": "http://127.0.0.1:7897", "https": "http://127.0.0.1:7897"})
        print(f"[DEBUG] HTTP 状态码: {response.status_code}")
        response.raise_for_status()

        soup = BeautifulSoup(response.text, 'html.parser')

        papers = []
        dl_list = soup.find('dl')

        if not dl_list:
            print("[!] 未找到 <dl> 标签，页面结构可能已变化")
            return

        entries = dl_list.find_all('dt')
        descriptions = dl_list.find_all('dd')

        print(f"[DEBUG] 找到 {len(entries)} 个 <dt> 标签和 {len(descriptions)} 个 <dd> 标签")

        for i, (entry, desc) in enumerate(zip(entries, descriptions)):
            # 暴力提取标题：优先从 <dd> 中找，因为标题通常在描述部分
            title = ""
            title_div = desc.find('div', class_='list-title')
            if title_div:
                title = title_div.get_text(strip=True).replace('Title:', '').strip()
            else:
                # 备用方案：直接从 <dd> 中找第一个有实质内容的文本
                all_text = desc.get_text(strip=True)
                if 'Title:' in all_text:
                    title = all_text.split('Title:')[1].split('Authors:')[0].strip()

            # 暴力提取摘要：直接从 <dd> 容器榨取所有文本
            abstract = desc.get_text(separator=' ', strip=True)
            # 清理干扰词
            abstract = abstract.replace('Abstract:', '').replace('▽ More', '').replace('△ Less', '')
            abstract = abstract.replace('Title:', '').replace('Authors:', '').replace('Comments:', '')
            abstract = abstract.replace('Subjects:', '').replace('Cite as:', '').strip()

            # 截取前 200 字符
            if len(abstract) > 200:
                abstract = abstract[:200] + "..."

            # 如果找到了标题和摘要就添加
            if title and abstract and len(abstract) > 20:  # 确保摘要有实质内容
                papers.append({
                    '序号': i + 1,
                    '论文标题': title,
                    '摘要': abstract
                })
            else:
                print(f"[DEBUG] 第 {i+1} 篇论文解析失败 - 标题: {'有' if title else '无'}, 摘要长度: {len(abstract) if abstract else 0}")

        if papers:
            df = pd.DataFrame(papers)
            file_path = 'university_courses_intel.xlsx'
            df.to_excel(file_path, index=False, engine='openpyxl')
            print(f"[✓] 成功抓取 {len(papers)} 篇论文，已导出至 {file_path}")

            # 美化 Excel
            beautify_excel(file_path, {
                '序号': 8,
                '论文标题': 40,
                '摘要': 80
            })
        else:
            print("[!] 未能解析到任何论文数据，请检查页面结构")

    except Exception as e:
        print(f"[✗] 任务失败: {e}")


def knowledge_harvest():
    """[2] 高分知识收割 - 抓取豆瓣读书 Top250"""
    print("\n[*] 启动高分知识收割...")

    url = "https://book.douban.com/top250"

    try:
        response = requests.get(url, headers=HEADERS, timeout=10, verify=False, proxies={"http": None, "https": None})
        response.raise_for_status()
        soup = BeautifulSoup(response.text, 'html.parser')

        books = []
        items = soup.find_all('tr', class_='item')

        for idx, item in enumerate(items[:25], 1):
            title_tag = item.find('div', class_='pl2').find('a')
            info_tag = item.find('p', class_='pl')
            rating_tag = item.find('span', class_='rating_nums')

            if title_tag and info_tag and rating_tag:
                title = title_tag.get('title', '').strip()
                author_info = info_tag.get_text(strip=True)
                rating = rating_tag.get_text(strip=True)

                books.append({
                    '排名': idx,
                    '书名': title,
                    '作者信息': author_info,
                    '评分': rating
                })

        df = pd.DataFrame(books)
        file_path = 'cognitive_improvement_intel.xlsx'
        df.to_excel(file_path, index=False, engine='openpyxl')
        print(f"[✓] 成功抓取 {len(books)} 本图书，已导出至 {file_path}")

        # 美化 Excel
        beautify_excel(file_path, {
            '排名': 8,
            '书名': 35,
            '作者信息': 50,
            '评分': 10
        })

    except Exception as e:
        print(f"[✗] 任务失败: {e}")


def entertainment_monitor():
    """[3] 赛博娱乐监控 - 抓取 Steam 热门特惠"""
    print("\n[*] 启动赛博娱乐监控...")

    url = "https://store.steampowered.com/search/?specials=1&filter=topsellers"

    try:
        response = requests.get(url, headers=HEADERS, timeout=10, verify=False, proxies={"http": "http://127.0.0.1:7897", "https": "http://127.0.0.1:7897"})
        response.raise_for_status()
        soup = BeautifulSoup(response.text, 'html.parser')

        games = []
        items = soup.find_all('a', class_='search_result_row')

        for idx, item in enumerate(items[:20], 1):
            title_tag = item.find('span', class_='title')
            price_div = item.find('div', class_='discount_prices')

            if title_tag and price_div:
                title = title_tag.get_text(strip=True)
                original_price = price_div.find('div', class_='discount_original_price')
                final_price = price_div.find('div', class_='discount_final_price')

                original = original_price.get_text(strip=True) if original_price else "N/A"
                final = final_price.get_text(strip=True) if final_price else "N/A"

                games.append({
                    '序号': idx,
                    '游戏名称': title,
                    '原价': original,
                    '折扣价': final
                })

        df = pd.DataFrame(games)
        file_path = 'entertainment_and_leisure_intel.xlsx'
        df.to_excel(file_path, index=False, engine='openpyxl')
        print(f"[✓] 成功抓取 {len(games)} 款游戏，已导出至 {file_path}")

        # 美化 Excel
        beautify_excel(file_path, {
            '序号': 8,
            '游戏名称': 40,
            '原价': 12,
            '折扣价': 12
        })

    except Exception as e:
        print(f"[✗] 任务失败: {e}")


def print_menu():
    """打印极客风格交互菜单"""
    print("\n" + "="*60)
    print("█████╗ ██╗   ██╗██████╗ ███████╗██████╗     ██╗███╗   ██╗████████╗███████╗██╗     ")
    print("██╔══██╗╚██╗ ██╔╝██╔══██╗██╔════╝██╔══██╗    ██║████╗  ██║╚══██╔══╝██╔════╝██║     ")
    print("██║  ╚═╝ ╚████╔╝ ██████╦╝█████╗  ██████╔╝    ██║██╔██╗ ██║   ██║   █████╗  ██║     ")
    print("██║  ██╗  ╚██╔╝  ██╔══██╗██╔══╝  ██╔══██╗    ██║██║╚██╗██║   ██║   ██╔══╝  ██║     ")
    print("╚█████╔╝   ██║   ██████╦╝███████╗██║  ██║    ██║██║ ╚████║   ██║   ███████╗███████╗")
    print(" ╚════╝    ╚═╝   ╚═════╝ ╚══════╝╚═╝  ╚═╝    ╚═╝╚═╝  ╚═══╝   ╚═╝   ╚══════╝╚══════╝")
    print("="*60)
    print("          全能赛博情报中心 v1.0 | Cyber Intelligence Hub")
    print("="*60)
    print("\n[任务列表]")
    print("  [1] 学术前沿雷达      -> arXiv 最新 CS 论文抓取")
    print("  [2] 高分知识收割      -> 豆瓣读书 Top250 数据采集")
    print("  [3] 赛博娱乐监控      -> Steam 热门特惠游戏追踪")
    print("  [0] 退出系统")
    print("="*60)


def main():
    """主程序入口"""
    while True:
        print_menu()
        choice = input("\n[>] 请输入任务编号: ").strip()

        if choice == '1':
            academic_radar()
        elif choice == '2':
            knowledge_harvest()
        elif choice == '3':
            entertainment_monitor()
        elif choice == '0':
            print("\n[*] 系统关闭中...")
            print("[✓] 感谢使用全能赛博情报中心！")
            break
        else:
            print("\n[!] 无效指令，请重新输入")

        time.sleep(1)


if __name__ == "__main__":
    main()
